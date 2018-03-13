import openpyxl as xl
import pandas as pd

from .. import utils

class XLSX(object):

  _fileName = None
  _xlBook = None
  _data = None

  def __init__(self, fileName):
    self._fileName = fileName
    self._xlBook = xl.load_workbook(fileName)
    self._data = {}
  #edef

  def loadSheet(self, sheetName, index=None, columns=None, header=False, names=None, skiprows=0, refresh=False):
    if sheetName not in self._xlBook.sheetnames:
      utils.error("Sheet '%s' not in workbook '%s'." % (sheetName, self._fileName))
      return None
    #fi

    sheetData = self._xlBook[sheetName].values
    sheetData = list(sheetData)[skiprows:]
    if header:
      columns = sheetData[0]
      sheetData = sheetData[1:]
    #fi

    self._data[sheetName] = pd.DataFrame(sheetData, columns=columns, index=index)
    return self._data[sheetName]
  #edef

  def isSheetLoaded(self, sheetName):
    return sheetName in self._data
  #edef

  def __getitem__(self, sheetName):
    if (sheetName in self._data):
      return self._data[sheetName]
    else:
      utils.error("You have not yet loaded the sheet '%s'" % sheetName)
      return None
    #fi
  #edef

  def __str__(self):
    dstr  = "XLSX object\n"
    dstr += " Where: %s\n" % (self._fileName if self._fileName is not None else hex(id(self)))
    dstr += " Sheets:\n"
    for sheet in self._xlBook.sheetnames:
      dstr += "  * [%s] %s\n" % ( 'X' if sheet in self._data else ' ', sheet)
    #efor
    return dstr
  #edef 

  def toTSV(self, sheetName, outFile, **kwargs):
    self[sheetName].to_csv(outFile, **kwargs)
  #edef

#eclass
